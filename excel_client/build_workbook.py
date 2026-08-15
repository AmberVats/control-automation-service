"""
Builds the pre-styled Excel client workbook for HSBC Product Control Analytics.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_control_runner_workbook(file_path="excel_client/ControlRunner.xlsx"):
    wb = Workbook()

    # Define color schemes (HSBC Corporate Theme: Crimson/Charcoal/Silver)
    header_fill = PatternFill(start_color="DB0011", end_color="DB0011", fill_type="solid")
    dark_fill = PatternFill(start_color="2D3142", end_color="2D3142", fill_type="solid")
    card_fill = PatternFill(start_color="F4F5F6", end_color="F4F5F6", fill_type="solid")
    accent_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    font_title = Font(name="Calibri", size=16, bold=True, color="2D3142")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="555555")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="2D3142")
    font_normal = Font(name="Calibri", size=11, color="000000")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # -------------------------------------------------------------
    # Sheet 1: Control Panel
    # -------------------------------------------------------------
    ws_panel = wb.active
    ws_panel.title = "Control Panel"
    ws_panel.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_panel["B2"] = "HSBC Product Control Analytics — Control Automation Service"
    ws_panel["B2"].font = font_title
    ws_panel["B3"] = "Citizen Developer Framework — REST API Microservice Interface (v1.0)"
    ws_panel["B3"].font = font_subtitle

    # Action Controls & Parameter Card
    ws_panel["B5"] = "Select Control:"
    ws_panel["B5"].font = font_bold
    ws_panel["C5"] = "eod_position_break"
    ws_panel["C5"].font = font_bold
    ws_panel["C5"].fill = accent_fill
    ws_panel["C5"].alignment = Alignment(horizontal="center")

    ws_panel["B6"] = "As-Of Date:"
    ws_panel["B6"].font = font_bold
    ws_panel["C6"] = "2026-08-15"
    ws_panel["C6"].alignment = Alignment(horizontal="center")

    # Execution KPI / Status Cards
    ws_panel["E5"] = "Latest Run Status:"
    ws_panel["E5"].font = font_bold
    ws_panel["F5"] = "READY"
    ws_panel["F5"].font = font_bold
    ws_panel["F5"].alignment = Alignment(horizontal="center")
    ws_panel["F5"].fill = accent_fill

    ws_panel["E6"] = "Breaches Detected:"
    ws_panel["E6"].font = font_bold
    ws_panel["F6"] = 0
    ws_panel["F6"].alignment = Alignment(horizontal="center")

    ws_panel["E7"] = "Duration:"
    ws_panel["E7"].font = font_bold
    ws_panel["F7"] = "—"
    ws_panel["F7"].alignment = Alignment(horizontal="center")

    # Instructions box
    ws_panel["B8"] = "Controls Catalogue (Registered Microservice Components):"
    ws_panel["B8"].font = font_bold

    headers_panel = ["Control Name", "Version", "Component", "Owner", "Schedule", "Status", "Config Hash"]
    for col_idx, h in enumerate(headers_panel, start=2):
        cell = ws_panel.cell(row=10, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_controls = [
        ("eod_position_break", "2", "reconciliation.two_way_match", "product_control_analytics", "0 18 * * 1-5", "ENABLED", "7a9b1c2d3e4f..."),
        ("market_price_tolerance", "1", "tolerance.threshold_check", "product_control_analytics", "30 17 * * 1-5", "ENABLED", "8b0c2d3e4f5a..."),
        ("trade_completeness", "1", "quality.completeness", "product_control_analytics", "0 19 * * 1-5", "ENABLED", "9c1d3e4f5a6b..."),
        ("trade_referential_integrity", "1", "quality.referential_integrity", "product_control_analytics", "15 19 * * 1-5", "ENABLED", "0d2e4f5a6b7c..."),
        ("market_feed_staleness", "1", "quality.staleness", "product_control_analytics", "45 17 * * 1-5", "ENABLED", "1e3f5a6b7c8d..."),
    ]

    for r_idx, row_data in enumerate(sample_controls, start=11):
        for c_idx, val in enumerate(row_data, start=2):
            c = ws_panel.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_normal
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if c_idx in (3, 6, 7, 8) else "left")

    # -------------------------------------------------------------
    # Sheet 2: Exceptions
    # -------------------------------------------------------------
    ws_exc = wb.create_sheet(title="Exceptions")
    ws_exc.views.sheetView[0].showGridLines = True

    exc_headers = ["Exception ID", "Run ID", "Exception Type", "Key Data", "Field", "Source Value", "Target Value", "Difference", "Message"]
    for col_idx, h in enumerate(exc_headers, start=1):
        cell = ws_exc.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center")

    sample_exceptions = [
        (1, "a1b2c3d4-e5f6-7890", "VALUE_MISMATCH", '["2026-08-15", "EQ_MSFT", "GLOBAL_EQ"]', "market_value", "960120.00", "960000.00", 120.00, "Market value difference $120.00 exceeds tolerance $50.00"),
        (2, "a1b2c3d4-e5f6-7890", "MISSING_TARGET", '["2026-08-15", "EQ_GOOG", "GLOBAL_EQ"]', "quantity", "1500", "None", 1500, "Position exists in Risk system but missing from Books & Records"),
    ]
    for r_idx, row_data in enumerate(sample_exceptions, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws_exc.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_normal
            c.border = thin_border

    # -------------------------------------------------------------
    # Sheet 3: Run History
    # -------------------------------------------------------------
    ws_hist = wb.create_sheet(title="Run History")
    ws_hist.views.sheetView[0].showGridLines = True

    hist_headers = ["Run ID", "Control Name", "Version", "Status", "Breaches", "Duration (ms)", "Triggered By", "Timestamp"]
    for col_idx, h in enumerate(hist_headers, start=1):
        cell = ws_hist.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center")

    sample_history = [
        ("a1b2c3d4-e5f6-7890", "eod_position_break", "2", "BREACH", 2, 42.5, "excel_vba_client", "2026-08-15 18:00:01"),
        ("f9e8d7c6-b5a4-3210", "market_price_tolerance", "1", "PASS", 0, 12.1, "scheduler", "2026-08-15 17:30:00"),
        ("12345678-abcd-ef01", "trade_completeness", "1", "PASS", 0, 18.7, "scheduler", "2026-08-15 19:00:00"),
    ]
    for r_idx, row_data in enumerate(sample_history, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws_hist.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_normal
            c.border = thin_border
            if c_idx == 4:
                c.font = Font(name="Calibri", bold=True, color="006100" if val == "PASS" else "9C0006")
                c.fill = pass_fill if val == "PASS" else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Auto-fit columns
    for ws in [ws_panel, ws_exc, ws_hist]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(file_path)
    print(f"Generated Excel client template at {file_path}")

if __name__ == "__main__":
    create_control_runner_workbook()
